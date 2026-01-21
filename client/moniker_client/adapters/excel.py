"""Excel adapter - reads Excel files."""

from __future__ import annotations

from pathlib import Path
from typing import Any, TYPE_CHECKING

if TYPE_CHECKING:
    from ..client import ResolvedSource
    from ..config import ClientConfig

from .base import BaseAdapter


class ExcelAdapter(BaseAdapter):
    """
    Adapter for Excel files (.xlsx, .xls).

    Reads Excel files from local filesystem or network paths.
    """

    def fetch(
        self,
        resolved: ResolvedSource,
        config: ClientConfig,
        **kwargs,
    ) -> Any:
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl required for Excel: pip install openpyxl")

        conn_info = resolved.connection
        params = resolved.params

        base_path = Path(conn_info.get("base_path", "."))
        file_path_pattern = resolved.query or "{path}.xlsx"

        # Build full file path
        file_path = base_path / file_path_pattern

        if not file_path.exists():
            from ..client import NotFoundError
            raise NotFoundError(f"Excel file not found: {file_path}")

        sheet_name = params.get("sheet")
        header_row = params.get("header_row", 1)

        # Read Excel file
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        try:
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    from ..client import NotFoundError
                    raise NotFoundError(f"Sheet '{sheet_name}' not found")
                ws = wb[sheet_name]
            else:
                ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []

            # Get headers
            header_idx = header_row - 1
            if header_idx >= len(rows):
                raise ValueError(f"Header row {header_row} not found")

            headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[header_idx])]

            # Build data
            data = []
            for row in rows[header_idx + 1:]:
                if any(cell is not None for cell in row):
                    record = dict(zip(headers, row))
                    data.append(record)

            return data

        finally:
            wb.close()

    def list_children(
        self,
        resolved: ResolvedSource,
        config: ClientConfig,
    ) -> list[str]:
        """List Excel files in directory."""
        conn_info = resolved.connection
        base_path = Path(conn_info.get("base_path", "."))

        if not base_path.is_dir():
            return []

        children = []
        for item in base_path.iterdir():
            if item.is_dir():
                children.append(item.name)
            elif item.suffix in (".xlsx", ".xls"):
                children.append(item.stem)

        return sorted(children)
