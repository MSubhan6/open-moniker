"""Excel file adapter."""

from __future__ import annotations

import time
from pathlib import Path
from typing import Any

from ..catalog.types import SourceBinding, SourceType
from ..moniker.types import Moniker
from .base import (
    AdapterError,
    AdapterNotFoundError,
    AdapterResult,
    DataAdapter,
)


class ExcelAdapter(DataAdapter):
    """
    Adapter for Excel files (.xlsx, .xls).

    Config:
        base_path: Base directory for files
        file_pattern: Pattern for file path (can include {path} placeholder)
        sheet: Sheet name or index (default: first sheet)
        header_row: Row number for headers (default: 0)
        skip_rows: Rows to skip at start
    """

    @property
    def source_type(self) -> SourceType:
        return SourceType.EXCEL

    async def fetch(
        self,
        moniker: Moniker,
        binding: SourceBinding,
        sub_path: str | None = None,
    ) -> AdapterResult:
        start = time.perf_counter()

        try:
            import openpyxl
        except ImportError:
            raise AdapterError("openpyxl required for Excel support: pip install openpyxl")

        base_path = Path(binding.config.get("base_path", "."))
        file_pattern = binding.config.get("file_pattern", "{path}.xlsx")
        sheet_name = binding.config.get("sheet")
        header_row = binding.config.get("header_row", 1)
        skip_rows = binding.config.get("skip_rows", 0)

        path_str = sub_path or str(moniker.path)
        file_path = base_path / file_pattern.format(path=path_str)

        if not file_path.exists():
            raise AdapterNotFoundError(f"Excel file not found: {file_path}")

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    raise AdapterNotFoundError(f"Sheet '{sheet_name}' not found")
                ws = wb[sheet_name]
            else:
                ws = wb.active

            rows = list(ws.iter_rows(min_row=skip_rows + 1, values_only=True))
            if not rows:
                return AdapterResult(data=[], source_type=self.source_type)

            # Get headers
            header_idx = header_row - skip_rows - 1
            if header_idx < 0 or header_idx >= len(rows):
                raise AdapterError(f"Header row {header_row} not found")

            headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[header_idx])]

            # Build data
            data = []
            for row in rows[header_idx + 1:]:
                if any(cell is not None for cell in row):  # Skip empty rows
                    record = dict(zip(headers, row))
                    data.append(record)

            wb.close()

        except AdapterError:
            raise
        except Exception as e:
            raise AdapterError(f"Failed to read Excel file {file_path}: {e}") from e

        elapsed = (time.perf_counter() - start) * 1000

        return AdapterResult(
            data=data,
            source_type=self.source_type,
            source_path=str(file_path),
            query_ms=elapsed,
            row_count=len(data),
            metadata={"sheet": sheet_name or "active", "headers": headers},
        )

    async def list_children(
        self,
        moniker: Moniker,
        binding: SourceBinding,
        sub_path: str | None = None,
    ) -> list[str]:
        base_path = Path(binding.config.get("base_path", "."))
        path_str = sub_path or str(moniker.path)

        dir_path = base_path / path_str
        if not dir_path.is_dir():
            return []

        children = []
        for item in dir_path.iterdir():
            if item.is_dir():
                children.append(item.name)
            elif item.suffix in (".xlsx", ".xls"):
                children.append(item.stem)

        return sorted(children)

    async def list_sheets(self, file_path: Path) -> list[str]:
        """List sheets in an Excel file."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets
        except Exception:
            return []
